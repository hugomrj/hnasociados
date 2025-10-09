from datetime import date
from calendar import month_name
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models.cliente_model import Cliente

VENCIMIENTOS = {
    '0': 7, '1': 9, '2': 11, '3': 13, '4': 15,
    '5': 17, '6': 19, '7': 21, '8': 23, '9': 25,
}

def calcular_vencimiento(cedula, anio, mes):
    if not cedula:
        return None
    ultimo = cedula[-1]
    dia = VENCIMIENTOS.get(ultimo)
    if not dia:
        return None
    try:
        return date(anio, mes, dia)
    except ValueError:
        return None

def calendario_perpetuo(request):
    hoy = date.today()
    anio = hoy.year
    mes = hoy.month
    
    # Filtros
    filtro_estado = request.GET.get("estado", "todos")
    dia_especifico = request.GET.get("dia", "")
    descargar = request.GET.get("descargar", "")

    clientes = Cliente.objects.all().order_by("nombre")
    data = []

    # Primero, calcular TODOS los días con vencimientos (sin filtros)
    todos_los_dias = set()
    for c in clientes:
        fecha_vto = calcular_vencimiento(c.cedula, anio, mes)
        if fecha_vto:
            todos_los_dias.add(fecha_vto.day)

    # Ordenar la lista de días
    dias_con_vencimientos = sorted(list(todos_los_dias))

    # Ahora aplicar los filtros a los datos
    for c in clientes:
        fecha_vto = calcular_vencimiento(c.cedula, anio, mes)
        if fecha_vto:
            dias_restantes = (fecha_vto - hoy).days
            
            # Aplicar filtro de estado
            if filtro_estado != "todos":
                if filtro_estado == "vencidos" and dias_restantes >= 0:
                    continue
                elif filtro_estado == "proximos" and (dias_restantes < 0 or dias_restantes > 7):
                    continue
                elif filtro_estado == "por_vencer" and (dias_restantes < 0 or dias_restantes <= 7):
                    continue
                elif filtro_estado == "pendientes" and dias_restantes < 0:
                    continue
            
            # Aplicar filtro de día específico si está seleccionado
            if dia_especifico and dia_especifico != "":
                if fecha_vto.day != int(dia_especifico):
                    continue
            
            data.append({
                "nombre": f"{c.nombre} {c.apellido}",
                "cedula": c.cedula,
                "terminacion": c.cedula[-1] if c.cedula else "",
                "vencimiento": fecha_vto,
                "vencimiento_str": fecha_vto.strftime("%d/%m/%Y"),
                "dias_restantes": dias_restantes,
                "esta_vencido": dias_restantes < 0,
                "esta_proximo": 0 <= dias_restantes <= 7,
                "dia_del_mes": fecha_vto.day,
            })

    # Ordenar por fecha de vencimiento
    data.sort(key=lambda x: x["vencimiento"])

    # Si se solicita descargar Excel
    if descargar == "excel":
        return generar_excel(data, mes, anio, filtro_estado, dia_especifico)

    # Calcular contadores para todos los clientes del mes (sin filtrar)
    todos_los_clientes = []
    for c in Cliente.objects.all():
        fecha_vto = calcular_vencimiento(c.cedula, anio, mes)
        if fecha_vto:
            dias_restantes = (fecha_vto - hoy).days
            todos_los_clientes.append({
                "dias_restantes": dias_restantes,
                "esta_vencido": dias_restantes < 0,
                "esta_proximo": 0 <= dias_restantes <= 7,
            })
    
    total_general = len(todos_los_clientes)
    vencidos_general = sum(1 for c in todos_los_clientes if c["esta_vencido"])
    proximos_general = sum(1 for c in todos_los_clientes if c["esta_proximo"])
    por_vencer_general = total_general - vencidos_general - proximos_general

    return render(request, "app/calendario/list.html", {
        "clientes": data,
        "mes_actual": mes,
        "anio_actual": anio,
        "filtro_estado": filtro_estado,
        "dia_seleccionado": dia_especifico,
        "dias_con_vencimientos": dias_con_vencimientos,  # Esta lista ahora es COMPLETA
        "total_clientes": len(data),
        "total_general": total_general,
        "vencidos_general": vencidos_general,
        "proximos_general": proximos_general,
        "por_vencer_general": por_vencer_general,
        "hoy": hoy,
    })

def generar_excel(data, mes, anio, filtro_estado, dia_especifico):
    # ... (el código de generar_excel se mantiene igual)
    # Crear el libro de trabajo y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = f"Vencimientos {mes}-{anio}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados
    headers = ['Cliente', 'Cédula/RUC', 'Terminación', 'Fecha Vencimiento', 'Días Restantes', 'Estado']
    ws.append(headers)
    
    # Aplicar estilo a los encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Llenar datos
    for item in data:
        estado = "Vencido" if item['esta_vencido'] else "Próximo" if item['esta_proximo'] else "Pendiente"
        ws.append([
            item['nombre'],
            item['cedula'],
            item['terminacion'],
            item['vencimiento_str'],
            item['dias_restantes'],
            estado
        ])
    
    # Ajustar anchos de columna
    column_widths = [40, 15, 12, 18, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con filtros aplicados
    filename = f"vencimientos_{mes}_{anio}"
    if filtro_estado != "todos":
        filename += f"_{filtro_estado}"
    if dia_especifico:
        filename += f"_dia{dia_especifico}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    return response